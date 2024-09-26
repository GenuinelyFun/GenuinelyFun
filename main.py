import customtkinter as ctk
from tkinter import filedialog, messagebox
import threading
import os
import subprocess
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import database_loop, database_panel, database_board, database_io_report, database_zone, database_address_report, database_control_groups_detailed

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

class JsonToExcelConverter:
    def __init__(self, root):
        self.root = root
        self.animating = False
        self.animation_step = 0
        self.root.title("Fire Expert Export Tool v24.03.15_02")
        self.root.minsize(900, 500)
        self.file_paths = {"json": "", "excel": ""}
        self.setup_ui()

    def setup_ui(self):
        self.setup_file_labels()
        self.setup_status_and_warning_labels()
        self.setup_file_selection_buttons()
        self.setup_process_and_open_excel_buttons()

    def setup_file_labels(self):
        self.file_frame = ctk.CTkFrame(self.root)
        self.file_frame.pack(padx=3, pady=3, expand=False, fill='x')
        self.file_labels = [
            ctk.CTkLabel(self.file_frame, text="No JSON file selected", anchor="center"),
            ctk.CTkLabel(self.file_frame, text="No Excel file selected", anchor="center")
        ]
        for label in self.file_labels:
            label.pack(padx=10, pady=1)

    def setup_status_and_warning_labels(self):
        status_texts = ["", "Select files to start the conversion process"]
        self.status_label = ctk.CTkLabel(self.root, text="\n".join(status_texts))
        self.status_label.pack(padx=10, pady=10)
        warning_text = "This tool can make mistakes. Consider checking important information."
        self.warning_label = ctk.CTkLabel(self.root, text=warning_text, font=("Helvetica", 11.5, "normal"))
        self.warning_label.pack(side="bottom", padx=10, pady=5)

    def setup_file_selection_buttons(self):
        self.file_select_buttons = [
            ctk.CTkButton(self.root, text="Select JSON File", command=lambda: self.select_file("json", filedialog.askopenfilename, [("JSON files", "*.json")])),
            ctk.CTkButton(self.root, text="Select Excel File", command=lambda: self.select_file("excel", filedialog.asksaveasfilename, [("Excel files", "*.xlsx")], ".xlsx"))
        ]
        for button in self.file_select_buttons:
            button.pack(padx=10, pady=5)

    def setup_process_and_open_excel_buttons(self):
        self.process_button = ctk.CTkButton(self.root, text="Process and Save Data", command=self.process_and_save_data)
        self.process_button.pack(padx=10, pady=5)
        self.open_excel_button = ctk.CTkButton(self.root, text="Open Excel File", command=self.open_excel_file, state=ctk.DISABLED)

    def select_file(self, type, dialog, filetypes, defaultextension=None):
        file_path = dialog(filetypes=filetypes, defaultextension=defaultextension)
        if file_path:
            self.file_paths[type] = file_path
            filename = os.path.basename(file_path)
            self.update_file_label(type, filename)

    def update_file_label(self, type, filename):
        index = 0 if type == "json" else 1
        self.file_labels[index].configure(text=f"Selected {type.upper()}: {filename}")

    def process_and_save_data(self):
        if not all(self.file_paths.values()):
            messagebox.showerror("Error", "Please select both JSON and Excel files.")
            return
        self.toggle_ui_elements(state=ctk.DISABLED)
        self.start_loading_animation()
        threading.Thread(target=self.process_data_in_thread).start()

    def start_loading_animation(self):
        self.animating = True
        self.animation_step = 0
        self.animate_status_label()

    def animate_status_label(self):
        if not self.animating:
            return
        animation_states = ["Processing.", "Processing..", "Processing...", "Processing"]
        self.status_label.configure(text=animation_states[self.animation_step % len(animation_states)])
        self.animation_step += 1
        self.root.after(500, self.animate_status_label)

    def process_data_in_thread(self):
        modules = [
        #    (database_panel, "Panel"),
        #    (database_loop, "Loop"),
        #    (database_zone, "Zone"),
        #    (database_board, "Board"),
        #    (database_io_report, "Input and Output Report"),
        #    (database_address_report, "Address Report"),
            (database_control_groups_detailed, "Control Groups"),
        ]
        success_messages = []
        try:
            for module, sheet_name in modules:
                result = module.process_json_to_excel(self.file_paths["json"], self.file_paths["excel"], sheet_name=sheet_name)
                message = f"{sheet_name} data processed successfully." if result.startswith("Data has been successfully written") else f"Error processing {sheet_name}."
                success_messages.append(message)
                if result.startswith("Data has been successfully written"):
                    # Convert columns to tables immediately after processing each module
                    self.convert_all_cells_to_table(self.file_paths["excel"], sheet_name)
            self.root.after_idle(self.update_status, "\n".join(success_messages))
        except Exception as e:
            self.update_status(f"An unexpected error occurred: {str(e)}")

    def update_status(self, message):
        def update():
            self.animating = False
            self.status_label.configure(text=message)
            self.toggle_ui_elements(state=ctk.NORMAL if "successfully" in message else ctk.DISABLED)
            if "successfully" in message:
                self.open_excel_button.pack(padx=10, pady=5)
                self.open_excel_button.configure(state=ctk.NORMAL)
            else:
                self.open_excel_button.pack_forget()
        self.root.after(0, update)

    def convert_all_cells_to_table(self, excel_file_path, sheet_name):
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb[sheet_name]
        # Find all cells with any value
        max_row = ws.max_row
        max_column = ws.max_column
        table_ref = f"A1:{openpyxl.utils.get_column_letter(max_column)}{max_row}"
        table = Table(displayName=f"{sheet_name.replace(' ', '')}Table", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                              showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        ws.add_table(table)
        wb.save(excel_file_path)

    def toggle_ui_elements(self, state):
        self.process_button.configure(state=state)
        for button in self.file_select_buttons:
            button.configure(state=state)
        self.open_excel_button.configure(state=state)

    def open_excel_file(self):
        try:
            os.startfile(self.file_paths["excel"])
        except AttributeError:
            subprocess.call(['open' if os.name == 'mac' else 'xdg-open', self.file_paths["excel"]], shell=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open Excel file. Please open it manually. {e}")

if __name__ == "__main__":
    root = ctk.CTk()
    app = JsonToExcelConverter(root)
    root.mainloop()
